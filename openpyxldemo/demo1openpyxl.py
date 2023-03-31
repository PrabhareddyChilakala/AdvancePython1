import openpyxl as xl
wb=xl.load_workbook("C://exceldata//openexcel.xlsx")
# sheet1=wb['Sheet1']
# sheet2=wb['Sheet2']
sheet3=wb['Sheet3']
# print(sheet1.max_row)
# print(sheet1.max_column)
# print(sheet1.cell(1,1).value)
# sheet1.cell(10,10).value="hcl"
# sheet2.cell(5,1).value='=sum(A1:A4)'
# wb.save("C://exceldata//openexcel.xlsx")
excel_data=[]
# for i in range(1,6):
#     excel_data.append(sheet1.cell(i,3).value)
# print(excel_data)
for i in range(1,4):
    first = []
    for j in range(1,5):
        first.append(sheet3.cell(j,i).value)
    excel_data.append(first)
print(excel_data)