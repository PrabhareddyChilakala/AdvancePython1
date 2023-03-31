import openpyxl as xl
wb=xl.load_workbook("C://exceldata//xltodict.xlsx")
sheet=wb['Sheet1']
dictionary={}
lis=[]
# for i in range(1,sheet1.max_column+1):
#     for j in range(1,sheet1.max_row+1):
#         dictionary(sheet1.cell((j),i).value)
for i in range(1,8):
    for j in range(2,7):
        dictionary.setdefault(sheet.cell(1, i).value)
        lis.append(sheet.cell(2,i).value)
# print(lis)
print(dictionary)
# l=dictionary.update(lis)
# print(l)